#!/usr/bin/env python3
"""Prepara los datos mensuales de Transferencias para la PWA estática.

Uso:
  python tools/process_monthly_data.py

Entradas:
  data/source/Base_Transferencias.xlsx
  data/source/monthly/*.csv

Salidas:
  data/manifest-data.json
  data/chunks/transferencias-AAAA-MM.json
  data/audit/*.json
  docs/file-audit.json
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - mensaje operativo
    raise SystemExit(
        "Falta openpyxl. Ejecuta: python -m pip install -r tools/requirements.txt"
    ) from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = PROJECT_ROOT / "data" / "source"
MONTHLY_ROOT = SOURCE_ROOT / "monthly"
BASE_PATH = SOURCE_ROOT / "Base_Transferencias.xlsx"
CHUNK_ROOT = PROJECT_ROOT / "data" / "chunks"
AUDIT_ROOT = PROJECT_ROOT / "data" / "audit"
MANIFEST_PATH = PROJECT_ROOT / "data" / "manifest-data.json"
FILE_AUDIT_PATH = PROJECT_ROOT / "docs" / "file-audit.json"

EXPECTED_HEADERS = [
    "Año",
    "Semana",
    "Dia",
    "Tipo Operación",
    "CeCo",
    "Tienda",
    "Ingrediente",
    "Unidad de Medida",
    "Proveedor",
    "Cantidad",
    "Costo Unitario",
    "Costo Total",
]

MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}
MONTH_LABELS = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

ROW_INDEX = {
    "year": 0,
    "week": 1,
    "date": 2,
    "ceco": 3,
    "store": 4,
    "ingredient": 5,
    "unit": 6,
    "provider": 7,
    "quantity": 8,
    "unitCost": 9,
    "totalCost": 10,
    "region": 11,
    "dm": 12,
    "monthOrigin": 13,
    "fileOrigin": 14,
    "sourceRow": 15,
    "cecoOriginal": 16,
    "cecoNormalized": 17,
    "directoryMatch": 18,
    "operationType": 19,
}


@dataclass
class MonthlyResult:
    file: str
    month: int
    year: int
    encoding: str
    delimiter: str
    source_rows: int = 0
    valid_rows: int = 0
    visible_rows: int = 0
    duplicate_rows: int = 0
    non_inventory_rows: int = 0
    non_inventory_amount: float = 0.0
    patrol_rows: int = 0
    patrol_amount: float = 0.0
    patrol_non_inventory_rows: int = 0
    unresolved_provider_rows: int = 0
    unresolved_provider_amount: float = 0.0
    unresolved_provider_non_inventory_rows: int = 0
    directory_matches: int = 0
    directory_missing: int = 0
    invalid_dates: int = 0
    date_conserved_rows: int = 0
    date_converted_rows: int = 0
    date_ambiguous_rows: int = 0
    week_mismatch_rows: int = 0
    date_pattern: str = ""
    week_system: str = ""
    future_dates: int = 0
    zero_quantity_rows: int = 0
    zero_total_rows: int = 0
    cost_math_mismatch_rows: int = 0
    sign_inconsistent_rows: int = 0
    quantity_total: float = 0.0
    amount_total: float = 0.0
    visible_quantity_total: float = 0.0
    visible_amount_total: float = 0.0
    min_date: str = ""
    max_date: str = ""
    elapsed_seconds: float = 0.0
    status: str = "Procesado"
    message: str = ""


@dataclass(frozen=True)
class DateResolution:
    original: str
    detected_format: str
    normalized: date | None
    calculated_year: int | None
    calculated_week: int | None
    validation_result: str
    reason: str
    category: str


def text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value if value is not None else "").strip())


def canonical(value: Any) -> str:
    return unicodedata.normalize("NFKC", text(value)).casefold()


def normalize_header(value: Any) -> str:
    value = unicodedata.normalize("NFD", text(value))
    return re.sub(r"[^a-z0-9]+", "", value.encode("ascii", "ignore").decode().lower())


def normalize_ceco(value: Any) -> str:
    raw = text(value)
    if re.fullmatch(r"\d+\.0", raw):
        raw = raw[:-2]
    if raw.isdigit() and len(raw) <= 5:
        return raw.zfill(5)
    return raw


def provider_ceco(value: Any) -> str:
    match = re.match(r"^(\d{4,6})\b", text(value))
    return normalize_ceco(match.group(1)) if match else ""


def number(value: Any) -> float:
    raw = text(value).replace("$", "").replace(",", "")
    if raw.startswith("(") and raw.endswith(")"):
        raw = f"-{raw[1:-1]}"
    if not raw:
        return 0.0
    return float(raw)


def integer(value: Any) -> int:
    return int(round(number(value)))


def detect_csv(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()[:65536]
    encoding = "utf-8-sig"
    for candidate in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            sample = raw.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("No se pudo identificar una codificación compatible.")
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","
    return encoding, delimiter


def month_from_filename(path: Path) -> int | None:
    normalized = canonical(path.stem)
    for name, month in MONTHS.items():
        if re.search(rf"(^|[^a-z]){re.escape(name)}([^a-z]|$)", normalized):
            return month
    match = re.search(r"(?:^|[^0-9])(0?[1-9]|1[0-2])(?:[^0-9]|$)", normalized)
    return int(match.group(1)) if match else None


def date_candidates(value: Any) -> list[tuple[str, date]]:
    raw = text(value)
    candidates: list[tuple[str, date]] = []
    for label, fmt in (
        ("DD/MM/YYYY", "%d/%m/%Y"),
        ("M/D/YYYY", "%m/%d/%Y"),
        ("YYYY-MM-DD", "%Y-%m-%d"),
        ("DD-MM-YYYY", "%d-%m-%Y"),
        ("M-D-YYYY", "%m-%d-%Y"),
    ):
        try:
            parsed = datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
        if (label, parsed) not in candidates:
            candidates.append((label, parsed))
    return candidates


def source_week(value: date, system: str) -> int:
    if system == "ISO":
        return value.isocalendar().week
    if system == "DOMINGO_0":
        return int(value.strftime("%U"))
    if system == "LUNES_0":
        return int(value.strftime("%W"))
    raise ValueError(f"Sistema de semana no reconocido: {system}")


def detect_file_date_rules(
    raw_rows: list[dict[str, Any]],
    header_map: dict[str, str],
    expected_month: int,
    expected_year: int,
) -> tuple[str, str]:
    format_counts: Counter[str] = Counter()
    week_matches: Counter[str] = Counter()
    for raw in raw_rows:
        try:
            row_year = integer(raw.get(header_map["Año"]))
            row_week = integer(raw.get(header_map["Semana"]))
        except (TypeError, ValueError):
            continue
        exact = [
            (label, parsed)
            for label, parsed in date_candidates(raw.get(header_map["Dia"]))
            if parsed.year == row_year == expected_year
            and parsed.month == expected_month
        ]
        unique_dates = {parsed for _, parsed in exact}
        if len(unique_dates) != 1:
            continue
        parsed = next(iter(unique_dates))
        labels = {label for label, value in exact if value == parsed}
        if len(labels) == 1:
            format_counts[next(iter(labels))] += 1
        for system in ("ISO", "DOMINGO_0", "LUNES_0"):
            if source_week(parsed, system) == row_week:
                week_matches[system] += 1
    if not week_matches:
        raise ValueError("No fue posible comprobar el sistema de semana con fechas no ambiguas.")
    highest = max(week_matches.values())
    systems = sorted(system for system, count in week_matches.items() if count == highest)
    if len(systems) != 1:
        raise ValueError(f"El sistema de semana no es inequívoco: {systems}.")
    predominant = format_counts.most_common(1)[0][0] if format_counts else ""
    return predominant, systems[0]


def resolve_date(
    value: Any,
    expected_month: int,
    expected_year: int,
    reported_year: int,
    reported_week: int,
    week_system: str,
    predominant_format: str,
) -> DateResolution:
    raw = text(value)
    candidates = date_candidates(raw)
    compatible = [
        (label, parsed)
        for label, parsed in candidates
        if parsed.year == reported_year == expected_year
        and parsed.month == expected_month
        and source_week(parsed, week_system) == reported_week
    ]
    unique_dates = {parsed for _, parsed in compatible}
    if len(unique_dates) == 1:
        parsed = next(iter(unique_dates))
        labels = [label for label, value in compatible if value == parsed]
        if len(set(labels)) > 1:
            detected = "Coincidente DD/MM/YYYY = M/D/YYYY"
            category = "Conservada"
            reason = (
                f"Ambas máscaras producen {parsed.strftime('%d/%m/%Y')} y coinciden "
                f"con Año {reported_year}, Semana {reported_week} ({week_system}) y mes del CSV."
            )
        else:
            detected = labels[0]
            category = "Corregida" if detected in {"M/D/YYYY", "M-D-YYYY"} else "Conservada"
            reason = (
                f"{detected} coincide con Año {reported_year}, Semana {reported_week} "
                f"({week_system}), mes del CSV y patrón predominante {predominant_format or 'comprobado'}."
            )
        return DateResolution(
            original=raw,
            detected_format=detected,
            normalized=parsed,
            calculated_year=parsed.year,
            calculated_week=source_week(parsed, week_system),
            validation_result="Válida",
            reason=reason,
            category=category,
        )
    reason = (
        "Ninguna interpretación coincide simultáneamente con Año, Semana y mes del CSV."
        if not unique_dates
        else "Más de una interpretación distinta coincide con Año, Semana y mes del CSV."
    )
    return DateResolution(
        original=raw,
        detected_format="Fecha ambigua" if candidates else "Formato inválido",
        normalized=None,
        calculated_year=None,
        calculated_week=None,
        validation_result="Pendiente de revisión",
        reason=reason,
        category="Ambigua",
    )


def parse_month_date(value: Any, expected_month: int, expected_year: int) -> date | None:
    """Compatibilidad para validaciones unitarias simples."""
    candidates = {
        parsed
        for _, parsed in date_candidates(value)
        if parsed.month == expected_month and parsed.year == expected_year
    }
    return next(iter(candidates)) if len(candidates) == 1 else None


def resolve_latest_period(
    accepted_results: list[MonthlyResult], today: date
) -> tuple[MonthlyResult | None, bool]:
    current = next(
        (
            item
            for item in accepted_results
            if item.year == today.year and item.month == today.month
        ),
        None,
    )
    if current:
        return current, False
    if not accepted_results:
        return None, True
    return max(accepted_results, key=lambda item: (item.year, item.month)), True


def exact_fingerprint(row: dict[str, Any]) -> str:
    payload = "\x1f".join(text(row.get(header)) for header in EXPECTED_HEADERS)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def load_master() -> tuple[dict[str, dict[str, str]], set[str], dict[str, Any]]:
    if not BASE_PATH.exists():
        raise FileNotFoundError(f"No existe {BASE_PATH.relative_to(PROJECT_ROOT)}")
    workbook = load_workbook(BASE_PATH, read_only=True, data_only=False)
    required = {"Base_Directorio", "Non Inventory"}
    missing = sorted(required.difference(workbook.sheetnames))
    if missing:
        raise ValueError(f"Faltan hojas requeridas: {', '.join(missing)}")

    directory: dict[str, dict[str, str]] = {}
    directory_duplicates: list[dict[str, Any]] = []
    ws = workbook["Base_Directorio"]
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers[:4] != ["CeCo", "Tienda", "Región", "DM"]:
        raise ValueError(f"Encabezados incompatibles en Base_Directorio: {headers[:4]}")
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        ceco = normalize_ceco(values[0])
        if not ceco:
            continue
        record = {
            "ceco": ceco,
            "tienda": text(values[1]),
            "region": text(values[2]),
            "dm": text(values[3]),
            "excelRow": excel_row,
        }
        if ceco in directory:
            directory_duplicates.append(
                {"ceco": ceco, "first": directory[ceco], "duplicate": record}
            )
        else:
            directory[ceco] = record

    non_inventory: set[str] = set()
    non_inventory_rows = 0
    non_inventory_duplicate_rows = 0
    ws = workbook["Non Inventory"]
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers[:2] != ["Ingrediente", "Categoría Inventario"]:
        raise ValueError(f"Encabezados incompatibles en Non Inventory: {headers[:2]}")
    for values in ws.iter_rows(min_row=2, values_only=True):
        ingredient = canonical(values[0])
        if not ingredient:
            continue
        non_inventory_rows += 1
        if ingredient in non_inventory:
            non_inventory_duplicate_rows += 1
        non_inventory.add(ingredient)

    metadata = {
        "workbook": BASE_PATH.name,
        "sheets": workbook.sheetnames,
        "directoryRows": len(directory),
        "directoryDuplicateKeys": len(directory_duplicates),
        "directoryDuplicates": directory_duplicates,
        "nonInventoryRows": non_inventory_rows,
        "nonInventoryUniqueKeys": len(non_inventory),
        "nonInventoryDuplicateRows": non_inventory_duplicate_rows,
    }
    return directory, non_inventory, metadata


def dictionary_encoder() -> tuple[dict[str, list[str]], Any]:
    dictionary_names = [
        "tienda",
        "ingrediente",
        "unidad",
        "proveedor",
        "region",
        "dm",
        "mes",
        "archivo",
        "operacion",
    ]
    dictionaries = {name: [] for name in dictionary_names}
    positions = {name: {} for name in dictionary_names}

    def encode(name: str, value: Any) -> int:
        cleaned = text(value)
        if cleaned not in positions[name]:
            positions[name][cleaned] = len(dictionaries[name])
            dictionaries[name].append(cleaned)
        return positions[name][cleaned]

    return dictionaries, encode


def audit_record(
    *,
    month_label: str,
    filename: str,
    source_row: int,
    iso_date: str,
    row: dict[str, Any],
    ceco_original: str,
    ceco_normalized: str,
    region: str,
    dm: str,
    issue_type: str,
    original_value: str,
    normalized_value: str,
    result: str,
    action: str,
    status: str,
    risk: str,
) -> dict[str, Any]:
    return {
        "Mes": month_label,
        "Archivo": filename,
        "Fila": source_row,
        "Fecha": iso_date,
        "Semana": integer(row.get("Semana")) if text(row.get("Semana")) else None,
        "Folio": "",
        "CeCo": ceco_original,
        "CeCo normalizado": ceco_normalized,
        "Artículo": text(row.get("Ingrediente")),
        "Tipo inconsistencia": issue_type,
        "Valor original": original_value,
        "Valor normalizado": normalized_value,
        "Resultado cruce": result,
        "Acción aplicada": action,
        "Estado": status,
        "Riesgo": risk,
        "Región": region,
        "DM": dm,
        "Tienda": text(row.get("Tienda")),
        "Proveedor": text(row.get("Proveedor")),
        "Cantidad": number(row.get("Cantidad")),
        "Importe": number(row.get("Costo Total")),
    }


def process() -> dict[str, Any]:
    started = time.perf_counter()
    today = date.today()
    CHUNK_ROOT.mkdir(parents=True, exist_ok=True)
    AUDIT_ROOT.mkdir(parents=True, exist_ok=True)
    directory, non_inventory, master_meta = load_master()
    dictionaries, encode = dictionary_encoder()

    files = sorted(
        MONTHLY_ROOT.glob("*.csv"),
        key=lambda path: (month_from_filename(path) or 99, canonical(path.name)),
    )
    if not files:
        raise FileNotFoundError("No se encontraron CSV mensuales en data/source/monthly.")

    seen_fingerprints: set[str] = set()
    monthly_results: list[MonthlyResult] = []
    all_exclusions: list[dict[str, Any]] = []
    all_hidden: list[dict[str, Any]] = []
    all_inconsistencies: list[dict[str, Any]] = []
    date_audit_groups: dict[tuple[Any, ...], dict[str, Any]] = {}
    chunks: list[dict[str, Any]] = []
    weeks: set[int] = set()
    date_values: list[date] = []
    processed_chunk_names: set[str] = set()

    for path in files:
        file_started = time.perf_counter()
        filename_month = month_from_filename(path)
        if filename_month is None:
            monthly_results.append(
                MonthlyResult(
                    file=path.name,
                    month=0,
                    year=0,
                    encoding="",
                    delimiter="",
                    status="Rechazado",
                    message="El nombre no contiene un mes reconocible.",
                )
            )
            continue

        encoding, delimiter = detect_csv(path)
        with path.open("r", encoding=encoding, newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            original_headers = [text(value) for value in (reader.fieldnames or [])]
            mapped_headers = {
                normalize_header(header): header for header in original_headers
            }
            missing_headers = [
                header
                for header in EXPECTED_HEADERS
                if normalize_header(header) not in mapped_headers
            ]
            extra_headers = [
                header
                for header in original_headers
                if normalize_header(header)
                not in {normalize_header(item) for item in EXPECTED_HEADERS}
            ]
            if missing_headers:
                monthly_results.append(
                    MonthlyResult(
                        file=path.name,
                        month=filename_month,
                        year=0,
                        encoding=encoding,
                        delimiter=delimiter,
                        status="Rechazado",
                        message=(
                            f"Columnas faltantes: {', '.join(missing_headers)}. "
                            f"Columnas adicionales: {', '.join(extra_headers) or 'ninguna'}."
                        ),
                    )
                )
                continue

            header_map = {
                expected: mapped_headers[normalize_header(expected)]
                for expected in EXPECTED_HEADERS
            }
            raw_rows = list(reader)
            if not raw_rows:
                monthly_results.append(
                    MonthlyResult(
                        file=path.name,
                        month=filename_month,
                        year=0,
                        encoding=encoding,
                        delimiter=delimiter,
                        status="Rechazado",
                        message="El CSV está vacío; los demás meses pueden procesarse.",
                    )
                )
                continue

        source_years = Counter()
        for raw in raw_rows:
            try:
                source_years[integer(raw.get(header_map["Año"]))] += 1
            except (TypeError, ValueError):
                pass
        expected_year = source_years.most_common(1)[0][0] if source_years else today.year
        predominant_format, week_system = detect_file_date_rules(
            raw_rows,
            header_map,
            filename_month,
            expected_year,
        )
        result = MonthlyResult(
            file=path.name,
            month=filename_month,
            year=expected_year,
            encoding=encoding,
            delimiter=delimiter,
            source_rows=len(raw_rows),
            date_pattern=predominant_format,
            week_system=week_system,
        )
        month_label = MONTH_LABELS[filename_month - 1]
        output_rows: list[list[Any]] = []
        valid_dates: list[date] = []

        for source_row, raw_source in enumerate(raw_rows, start=2):
            row = {
                expected: raw_source.get(actual)
                for expected, actual in header_map.items()
            }
            ceco_original = text(row["CeCo"])
            ceco = normalize_ceco(ceco_original)
            directory_item = directory.get(ceco)
            region = directory_item["region"] if directory_item else ""
            dm = directory_item["dm"] if directory_item else ""
            try:
                week = integer(row["Semana"])
                year = integer(row["Año"])
            except (TypeError, ValueError) as exc:
                result.invalid_dates += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date="",
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Año o Semana inválidos",
                        original_value=f"Año {text(row['Año'])}; Semana {text(row['Semana'])}",
                        normalized_value="",
                        result="No procesado",
                        action="Registro separado para revisión",
                        status="Pendiente",
                        risk="Alto",
                    )
                )
                continue
            resolution = resolve_date(
                row["Dia"],
                filename_month,
                expected_year,
                year,
                week,
                week_system,
                predominant_format,
            )
            parsed = resolution.normalized
            date_key = (
                path.name,
                resolution.original,
                week,
                resolution.detected_format,
                parsed.isoformat() if parsed else "",
                resolution.validation_result,
                resolution.reason,
                resolution.category,
            )
            date_entry = date_audit_groups.setdefault(
                date_key,
                {
                    "Archivo": path.name,
                    "Mes CSV": month_label,
                    "Valor original": resolution.original,
                    "Formato detectado": resolution.detected_format,
                    "Fecha normalizada": parsed.strftime("%d/%m/%Y") if parsed else "",
                    "Año informado": year,
                    "Semana informada": week,
                    "Año calculado": resolution.calculated_year,
                    "Semana calculada": resolution.calculated_week,
                    "Sistema semana": week_system,
                    "Patrón predominante": predominant_format,
                    "Resultado validación": resolution.validation_result,
                    "Motivo": resolution.reason,
                    "Clasificación": resolution.category,
                    "Cantidad registros": 0,
                    "Primera fila": source_row,
                    "Última fila": source_row,
                },
            )
            date_entry["Cantidad registros"] += 1
            date_entry["Última fila"] = source_row
            if resolution.category == "Conservada":
                result.date_conserved_rows += 1
            elif resolution.category == "Corregida":
                result.date_converted_rows += 1
            else:
                result.date_ambiguous_rows += 1
            if parsed is None:
                result.invalid_dates += 1
                if any(
                    candidate.year == year
                    and candidate.month == filename_month
                    and source_week(candidate, week_system) != week
                    for _, candidate in date_candidates(row["Dia"])
                ):
                    result.week_mismatch_rows += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date="",
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Fecha ambigua o inconsistente con Año y Semana",
                        original_value=text(row["Dia"]),
                        normalized_value="",
                        result=resolution.validation_result,
                        action="Registro separado para revisión",
                        status="Pendiente",
                        risk="Alto",
                    )
                )
                continue
            result.valid_rows += 1
            valid_dates.append(parsed)
            date_values.append(parsed)
            if parsed > today:
                result.future_dates += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Fecha futura",
                        original_value=text(row["Dia"]),
                        normalized_value=parsed.isoformat(),
                        result="Marcado para revisión",
                        action="Excluido de Última actualización",
                        status="Pendiente",
                        risk="Alto",
                    )
                )

            fingerprint = exact_fingerprint(row)
            if fingerprint in seen_fingerprints:
                result.duplicate_rows += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Registro duplicado exacto",
                        original_value=fingerprint[:16],
                        normalized_value=fingerprint[:16],
                        result="Duplicado confirmado",
                        action="No consolidado; original conservado en CSV",
                        status="Controlado",
                        risk="Medio",
                    )
                )
                continue
            seen_fingerprints.add(fingerprint)

            try:
                quantity = number(row["Cantidad"])
                unit_cost = number(row["Costo Unitario"])
                total_cost = number(row["Costo Total"])
            except (TypeError, ValueError) as exc:
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Tipo de dato numérico inválido",
                        original_value=str(exc),
                        normalized_value="",
                        result="No procesado",
                        action="Registro separado para revisión",
                        status="Pendiente",
                        risk="Alto",
                    )
                )
                continue

            result.quantity_total += quantity
            result.amount_total += total_cost
            weeks.add(week)
            if abs(quantity) <= 1e-12:
                result.zero_quantity_rows += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Cantidad cero",
                        original_value=text(row["Cantidad"]),
                        normalized_value=str(quantity),
                        result="Requiere validación de origen",
                        action="Conservado; no se inventa corrección",
                        status="Pendiente",
                        risk="Medio",
                    )
                )
            if abs(total_cost) <= 1e-12:
                result.zero_total_rows += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Costo total cero",
                        original_value=text(row["Costo Total"]),
                        normalized_value=str(total_cost),
                        result="Requiere validación de origen",
                        action="Conservado; no se inventa corrección",
                        status="Pendiente",
                        risk="Medio",
                    )
                )
            # El CSV redondea el costo unitario a dos decimales. La tolerancia
            # absorbe ese redondeo por unidad y conserva solo diferencias que
            # no pueden explicarse por la precisión publicada.
            rounding_tolerance = max(0.05, abs(quantity) * 0.005 + 0.011)
            if (
                abs(quantity) > 1e-12
                and abs(unit_cost) > 1e-12
                and abs(abs(quantity * unit_cost) - abs(total_cost))
                > rounding_tolerance
            ):
                result.cost_math_mismatch_rows += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Cantidad × costo unitario no concilia",
                        original_value=(
                            f"{text(row['Cantidad'])} × {text(row['Costo Unitario'])}"
                            f" ≠ {text(row['Costo Total'])}"
                        ),
                        normalized_value=(
                            f"Diferencia {abs(abs(quantity * unit_cost) - abs(total_cost)):.2f}"
                        ),
                        result="Diferencia superior al redondeo publicado",
                        action="Conservado; revisar contra fuente de mayor precisión",
                        status="Pendiente",
                        risk="Medio",
                    )
                )
            if (
                abs(quantity) > 1e-12
                and abs(total_cost) > 1e-12
                and (quantity > 0) != (total_cost > 0)
            ):
                result.sign_inconsistent_rows += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Signo inconsistente",
                        original_value=(
                            f"Cantidad {text(row['Cantidad'])}; "
                            f"Costo Total {text(row['Costo Total'])}"
                        ),
                        normalized_value="",
                        result="Cantidad e importe tienen signos opuestos",
                        action="Conservado; requiere revisión manual",
                        status="Pendiente",
                        risk="Alto",
                    )
                )

            if directory_item:
                result.directory_matches += 1
            else:
                result.directory_missing += 1
                all_inconsistencies.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region="",
                        dm="",
                        issue_type="CeCo sin correspondencia",
                        original_value=ceco_original,
                        normalized_value=ceco,
                        result="Sin Región ni DM",
                        action="Conservado para revisión manual",
                        status="Pendiente",
                        risk="Alto",
                    )
                )

            provider = text(row["Proveedor"])
            is_patrol = bool(
                re.search(r"coffee[_\s-]*patrol|\bpatrol\b", provider, re.I)
                or provider.startswith("38100")
            )
            unresolved_provider = not provider_ceco(provider)
            if is_patrol:
                result.patrol_rows += 1
                result.patrol_amount += total_cost
            if unresolved_provider:
                result.unresolved_provider_rows += 1
                result.unresolved_provider_amount += total_cost

            if canonical(row["Ingrediente"]) in non_inventory:
                result.non_inventory_rows += 1
                result.non_inventory_amount += total_cost
                if is_patrol:
                    result.patrol_non_inventory_rows += 1
                if unresolved_provider:
                    result.unresolved_provider_non_inventory_rows += 1
                excluded = audit_record(
                    month_label=month_label,
                    filename=path.name,
                    source_row=source_row,
                    iso_date=parsed.isoformat(),
                    row=row,
                    ceco_original=ceco_original,
                    ceco_normalized=ceco,
                    region=region,
                    dm=dm,
                    issue_type="Artículo Non Inventory",
                    original_value=text(row["Ingrediente"]),
                    normalized_value=text(row["Ingrediente"]),
                    result="Coincidencia exacta normalizada",
                    action="Excluido de indicadores; original conservado",
                    status="Controlado",
                    risk="Bajo",
                )
                excluded["Motivo exclusión"] = "Coincide con Base_Transferencias.xlsx / Non Inventory"
                all_exclusions.append(excluded)
                continue

            if is_patrol:
                all_hidden.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Patrol",
                        original_value=provider,
                        normalized_value=text(provider),
                        result="Oculto de la vista operativa",
                        action="Conservado en registro técnico",
                        status="Controlado",
                        risk="Bajo",
                    )
                )
            if unresolved_provider:
                all_hidden.append(
                    audit_record(
                        month_label=month_label,
                        filename=path.name,
                        source_row=source_row,
                        iso_date=parsed.isoformat(),
                        row=row,
                        ceco_original=ceco_original,
                        ceco_normalized=ceco,
                        region=region,
                        dm=dm,
                        issue_type="Proveedor sin CeCo / Transferencias",
                        original_value=provider,
                        normalized_value=text(provider),
                        result="Oculto de la vista operativa",
                        action="Conservado en registro técnico",
                        status="Controlado",
                        risk="Medio",
                    )
                )

            result.visible_rows += 1
            result.visible_quantity_total += quantity
            result.visible_amount_total += total_cost
            output_rows.append(
                [
                    year,
                    week,
                    parsed.isoformat(),
                    ceco,
                    encode("tienda", row["Tienda"]),
                    encode("ingrediente", row["Ingrediente"]),
                    encode("unidad", row["Unidad de Medida"]),
                    encode("proveedor", provider),
                    quantity,
                    unit_cost,
                    total_cost,
                    encode("region", region),
                    encode("dm", dm),
                    encode("mes", month_label),
                    encode("archivo", path.name),
                    source_row,
                    ceco_original,
                    ceco,
                    bool(directory_item),
                    encode("operacion", row["Tipo Operación"]),
                ]
            )

        if valid_dates:
            result.min_date = min(valid_dates).isoformat()
            result.max_date = max(valid_dates).isoformat()
        result.elapsed_seconds = round(time.perf_counter() - file_started, 4)
        chunk_id = f"{expected_year}-{filename_month:02d}"
        chunk_path = CHUNK_ROOT / f"transferencias-{chunk_id}.json"
        chunk_payload = {
            "id": chunk_id,
            "sourceFile": path.name,
            "monthOrigin": month_label,
            "dateRange": {"min": result.min_date, "max": result.max_date},
            "rows": output_rows,
        }
        chunk_path.write_text(
            json.dumps(chunk_payload, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        processed_chunk_names.add(chunk_path.name)
        chunks.append(
            {
                "id": chunk_id,
                "year": expected_year,
                "month": filename_month,
                "label": month_label,
                "sourceFile": path.name,
                "weeks": sorted({row[ROW_INDEX["week"]] for row in output_rows}),
                "dateRange": {"min": result.min_date, "max": result.max_date},
                "records": len(output_rows),
                "sourceRecords": result.source_rows,
                "excludedRecords": result.non_inventory_rows,
                "hiddenPatrolRecords": result.patrol_rows,
                "hiddenUnresolvedProviderRecords": result.unresolved_provider_rows,
                "path": f"data/chunks/{chunk_path.name}",
            }
        )
        monthly_results.append(result)

    for old_chunk in CHUNK_ROOT.glob("transferencias-*.json"):
        if old_chunk.name not in processed_chunk_names:
            old_chunk.unlink()

    chunks.sort(key=lambda item: (item["year"], item["month"]))
    accepted_results = [item for item in monthly_results if item.status == "Procesado"]
    latest_period, fallback_used = resolve_latest_period(accepted_results, today)
    latest_valid_date = ""
    if latest_period and latest_period.max_date:
        candidates = [
            parsed
            for parsed in date_values
            if parsed.year == latest_period.year
            and parsed.month == latest_period.month
            and parsed <= today
        ]
        latest_valid_date = max(candidates).isoformat() if candidates else ""

    directory_rows = [
        {
            "ceco": item["ceco"],
            "tienda": item["tienda"],
            "region": item["region"],
            "dm": item["dm"],
        }
        for item in directory.values()
    ]
    directory_rows.sort(key=lambda item: item["ceco"])
    ceco_stores = {item["ceco"]: item["tienda"] for item in directory_rows}
    totals = {
        "sourceFiles": len(files),
        "processedFiles": len(accepted_results),
        "rejectedFiles": len(monthly_results) - len(accepted_results),
        "sourceRows": sum(item.source_rows for item in accepted_results),
        "validRows": sum(item.valid_rows for item in accepted_results),
        "recordsAuditoria": sum(item.visible_rows for item in accepted_results),
        "duplicateRows": sum(item.duplicate_rows for item in accepted_results),
        "nonInventoryRows": sum(item.non_inventory_rows for item in accepted_results),
        "nonInventoryAmount": round(
            sum(item.non_inventory_amount for item in accepted_results), 2
        ),
        "hiddenPatrolRows": sum(item.patrol_rows for item in accepted_results),
        "hiddenPatrolRowsInVisibleData": sum(
            item.patrol_rows - item.patrol_non_inventory_rows
            for item in accepted_results
        ),
        "hiddenPatrolAmount": round(
            sum(item.patrol_amount for item in accepted_results), 2
        ),
        "hiddenUnresolvedProviderRows": sum(
            item.unresolved_provider_rows for item in accepted_results
        ),
        "hiddenUnresolvedProviderRowsInVisibleData": sum(
            item.unresolved_provider_rows - item.unresolved_provider_non_inventory_rows
            for item in accepted_results
        ),
        "hiddenUnresolvedProviderAmount": round(
            sum(item.unresolved_provider_amount for item in accepted_results), 2
        ),
        "directoryMatches": sum(
            item.directory_matches for item in accepted_results
        ),
        "directoryMissing": sum(
            item.directory_missing for item in accepted_results
        ),
        "invalidDateRows": sum(item.invalid_dates for item in accepted_results),
        "dateConservedRows": sum(
            item.date_conserved_rows for item in accepted_results
        ),
        "dateConvertedRows": sum(
            item.date_converted_rows for item in accepted_results
        ),
        "dateAmbiguousRows": sum(
            item.date_ambiguous_rows for item in accepted_results
        ),
        "weekMismatchRows": sum(
            item.week_mismatch_rows for item in accepted_results
        ),
        "futureDateRows": sum(item.future_dates for item in accepted_results),
        "zeroTotalRows": sum(item.zero_total_rows for item in accepted_results),
        "zeroQuantityRows": sum(
            item.zero_quantity_rows for item in accepted_results
        ),
        "costMathMismatchRows": sum(
            item.cost_math_mismatch_rows for item in accepted_results
        ),
        "signInconsistentRows": sum(
            item.sign_inconsistent_rows for item in accepted_results
        ),
        "quantitySource": round(
            sum(item.quantity_total for item in accepted_results), 4
        ),
        "amountSource": round(
            sum(item.amount_total for item in accepted_results), 2
        ),
        "quantityVisible": round(
            sum(item.visible_quantity_total for item in accepted_results), 4
        ),
        "amountVisible": round(
            sum(item.visible_amount_total for item in accepted_results), 2
        ),
        "minDate": min(
            (item.min_date for item in accepted_results if item.min_date),
            default="",
        ),
        "maxDate": max(
            (item.max_date for item in accepted_results if item.max_date),
            default="",
        ),
    }
    totals["reconciliation"] = {
        "source": totals["sourceRows"],
        "visible": (
            totals["recordsAuditoria"]
            - totals["hiddenPatrolRowsInVisibleData"]
            - totals["hiddenUnresolvedProviderRowsInVisibleData"]
        ),
        "nonInventory": totals["nonInventoryRows"],
        "hiddenPatrol": totals["hiddenPatrolRowsInVisibleData"],
        "hiddenUnresolvedProvider": totals[
            "hiddenUnresolvedProviderRowsInVisibleData"
        ],
        "incidencesPending": totals["duplicateRows"] + totals["invalidDateRows"],
        "difference": (
            totals["sourceRows"]
            - (
                totals["recordsAuditoria"]
                - totals["hiddenPatrolRowsInVisibleData"]
                - totals["hiddenUnresolvedProviderRowsInVisibleData"]
            )
            - totals["nonInventoryRows"]
            - totals["hiddenPatrolRowsInVisibleData"]
            - totals["hiddenUnresolvedProviderRowsInVisibleData"]
            - totals["duplicateRows"]
            - totals["invalidDateRows"]
        ),
    }

    generation_elapsed = round(time.perf_counter() - started, 4)
    monthly_payload = [asdict(item) for item in monthly_results]
    audit_summary = {
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "processingSeconds": generation_elapsed,
        "master": master_meta,
        "monthly": monthly_payload,
        "totals": totals,
        "latestUpdate": {
            "date": latest_valid_date,
            "year": latest_period.year if latest_period else None,
            "month": latest_period.month if latest_period else None,
            "sourceFile": latest_period.file if latest_period else "",
            "currentMonthFileFound": not fallback_used,
            "fallbackUsed": fallback_used,
        },
    }
    (AUDIT_ROOT / "resumen-procesamiento.json").write_text(
        json.dumps(audit_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (AUDIT_ROOT / "exclusiones.json").write_text(
        json.dumps(all_exclusions, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (AUDIT_ROOT / "inconsistencias-datos.json").write_text(
        json.dumps(all_inconsistencies, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (AUDIT_ROOT / "fechas.json").write_text(
        json.dumps(
            sorted(
                date_audit_groups.values(),
                key=lambda item: (
                    item["Archivo"],
                    item["Primera fila"],
                    item["Valor original"],
                ),
            ),
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        encoding="utf-8",
    )
    (AUDIT_ROOT / "ocultos.json").write_text(
        json.dumps(all_hidden, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    manifest = {
        "version": f"v14-fechas-auditadas-{latest_valid_date or 'sin-fecha'}",
        "source": "CSV mensuales + Base_Transferencias.xlsx",
        "generatedAt": audit_summary["generatedAt"],
        "generatedFromHeaders": EXPECTED_HEADERS,
        "latestUpdate": latest_valid_date,
        "latestUpdateEvidence": audit_summary["latestUpdate"],
        "idx": ROW_INDEX,
        "dicts": dictionaries,
        "directory": directory_rows,
        "cecoStores": ceco_stores,
        "months": [
            {
                "value": item.month,
                "label": MONTH_LABELS[item.month - 1],
                "year": item.year,
                "records": item.visible_rows,
                "sourceRecords": item.source_rows,
                "sourceFile": item.file,
            }
            for item in accepted_results
        ],
        "weeks": sorted(weeks),
        "chunks": chunks,
        "totals": totals,
        "monthlyAudit": monthly_payload,
        "auditFiles": {
            "summary": "data/audit/resumen-procesamiento.json",
            "exclusions": "data/audit/exclusiones.json",
            "inconsistencies": "data/audit/inconsistencias-datos.json",
            "hidden": "data/audit/ocultos.json",
            "dates": "data/audit/fechas.json",
        },
        "businessRules": {
            "transferGroup": "Una transferencia por fecha, CeCo origen y CeCo destino.",
            "source": "Archivos CSV mensuales detectados por nombre y fecha interna.",
            "directoryUse": "Base_Directorio es la fuente oficial de Tienda, Región y DM.",
            "nonInventory": "Coincidencia exacta tras normalizar espacios y mayúsculas/minúsculas.",
            "match": "Ingrediente exacto; equivalencia económica solo cuando es inequívoca.",
            "surplusEntry": "El remanente de una entrada parcialmente utilizada no crea una transferencia duplicada.",
            "unresolvedProvider": "Proveedor sin CeCo / Transferencias se conserva en auditoría técnica y se oculta en la vista.",
            "patrol": "Patrol se conserva en auditoría técnica y se oculta en la vista.",
            "futureOnly": "Una entrada se asocia a una salida de la misma fecha o posterior.",
            "dateValidation": "Cada fecha se valida contra Año, semana ISO comprobada, mes del CSV y patrón predominante del archivo.",
            "sameDayIncluded": True,
            "costTolerance": 10,
            "duplicates": "Solo se retiran duplicados exactos de los 12 campos originales.",
        },
    }
    MANIFEST_PATH.write_text(
        json.dumps(manifest, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    build_file_audit()
    return audit_summary


def build_file_audit() -> None:
    paths = [
        path
        for path in PROJECT_ROOT.rglob("*")
        if path.is_file()
        and ".git" not in path.parts
        and "__pycache__" not in path.parts
        and ".DS_Store" not in path.name
    ]
    payload = {
        "version": "v14-fechas-auditadas",
        "limitBytes": 20 * 1024 * 1024,
        "files": [],
    }
    for path in sorted(paths):
        if path == FILE_AUDIT_PATH:
            continue
        raw = path.read_bytes()
        payload["files"].append(
            {
                "path": path.relative_to(PROJECT_ROOT).as_posix(),
                "bytes": len(raw),
                "mb": round(len(raw) / 1024 / 1024, 4),
                "over20MB": len(raw) > 20 * 1024 * 1024,
                "sha256": hashlib.sha256(raw).hexdigest()[:16],
            }
        )
    FILE_AUDIT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    try:
        summary = process()
    except Exception as error:  # pragma: no cover - salida operativa
        print(f"ERROR: {error}", file=sys.stderr)
        raise
    print(
        json.dumps(
            {
                "status": "ok",
                "months": len(summary["monthly"]),
                "rows": summary["totals"]["sourceRows"],
                "visible": summary["totals"]["recordsAuditoria"],
                "latestUpdate": summary["latestUpdate"],
                "processingSeconds": summary["processingSeconds"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
